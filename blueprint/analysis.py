import os
import json
from flask import Blueprint, request, render_template
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from sqlalchemy import JSON
from flask import send_from_directory

import config
from exts import db

analysis_bp = Blueprint('analysis', __name__, url_prefix='/a')
os.makedirs(config.UPLOAD_FOLDER, exist_ok=True)


# ----------- 模型 -------------
class QuestionaireModel(db.Model):
    '''
    问卷列表模型
    '''
    __tablename__ = 'q_list'
    __table_args__ = {'extend_existing': True}
    id = db.Column(db.Integer, primary_key=True, autoincrement=True)
    area = db.Column(db.String(10))
    code = db.Column(db.String(10))
    questions = db.Column(JSON)


# ----------- 工具函数 -------------
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in config.ALLOWED_EXTENSIONS


def extract_questionaire_code(sheet, remark_col_index):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        remark = row[remark_col_index - 1]
        if not remark:
            continue
        try:
            remark_data = json.loads(remark)
            if isinstance(remark_data, dict):
                print('获取问卷代码')
                print(f'remark_data的类型：{type(remark_data)}')
                return remark_data.get('c')
        except (json.JSONDecodeError, TypeError):
            continue
    return None


def write_answers_to_sheet(sheet, remark_col_index):
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        remark_cell = sheet.cell(row=i, column=remark_col_index)
        remark = remark_cell.value
        if not remark:
            continue

        try:
            remark_data = json.loads(remark)
            if isinstance(remark_data, dict):
                print('即将写入单元格')
                print(f'remark_data的类型：{type(remark_data)}')
                answers = remark_data.get('analysis', [])
                for j, value in enumerate(answers):
                    target_col = remark_col_index + j + 1
                    sheet.cell(row=i, column=target_col, value=value)
        except json.JSONDecodeError:
            continue


def write_question_titles(sheet, remark_col_index, questions):
    titles = [q.get('title') for q in questions]
    for i, title in enumerate(titles):
        sheet.cell(row=1, column=remark_col_index + i + 1, value=title)


# ----------- 路由入口 -------------
@analysis_bp.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'GET':
        return render_template('analysis/upload.html')

    # POST 请求处理
    file = request.files.get('file')
    if not file or file.filename == '':
        return render_template('error/400.html', error="文件为空或未选择文件")
    if not allowed_file(file.filename):
        return render_template('error/400.html', error="仅支持 .xlsx 文件")

    try:
        # 保存上传文件
        filename = secure_filename(file.filename)
        filepath = os.path.join(config.UPLOAD_FOLDER, filename)
        file.save(filepath)

        # 加载 Excel
        workbook = load_workbook(filepath, data_only=True)
        sheet_name = '企业拜访明细'
        sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.active

        if not sheet:
            return render_template("error/400.html", error="上传的 Excel 中未找到正确的 Sheet")

        header = [cell.value for cell in sheet[1]]
        if '备注' not in header:
            return render_template("error/400.html", error="未找到“备注”列，您上传的excel文件是易查查导出的excel吗？")

        remark_col_index = header.index('备注') + 1

        # 提取 code -> 查询问卷模板
        code = extract_questionaire_code(sheet, remark_col_index)
        if not code:
            return render_template("error/400.html", error="未找到问卷code，请检查上传文件是否是易查查导出的excel表格")

        questionaire = QuestionaireModel.query.filter_by(code=code).first()
        if not questionaire:
            return render_template("error/400.html", error=f"问卷 code [{code}] 未在系统中找到")

        questions = questionaire.questions

        # 写入 titles & answers
        write_question_titles(sheet, remark_col_index, questions)
        write_answers_to_sheet(sheet, remark_col_index)

        workbook.save(filepath)

        return render_template('analysis/success.html', filename=filename)

    except Exception as e:
        return render_template("error/500.html", e=str(e))


@analysis_bp.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(config.UPLOAD_FOLDER, filename, as_attachment=True)
